from __future__ import annotations

from pathlib import Path

from .schema import RawOperationRow

TARGET_SHEETS = ["GLOWNA", "ROB_MAT", "Pakowanie"]


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float))


def parse_legacy_excel(path: str | Path) -> dict:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError("openpyxl is required for legacy Excel import") from exc

    workbook = load_workbook(filename=path, data_only=True)
    found_sheets = [name for name in TARGET_SHEETS if name in workbook.sheetnames]

    operations: list[RawOperationRow] = []
    material_detected = False
    packaging_detected = False
    cost_summary_detected = False

    for sheet_name in found_sheets:
        ws = workbook[sheet_name]
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        header_map = {h.lower(): i for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            cells = ["" if v is None else v for v in row]
            as_text = " ".join(str(v).lower() for v in cells)

            if any(x in as_text for x in ["material", "materiał"]):
                material_detected = True
            if "pak" in as_text:
                packaging_detected = True
            if any(x in as_text for x in ["podsum", "summary", "razem", "total"]):
                cost_summary_detected = True

            op_name = None
            for key in ["operacja", "operation", "nazwa operacji"]:
                idx = header_map.get(key)
                if idx is not None and idx < len(cells) and cells[idx]:
                    op_name = str(cells[idx]).strip()
                    break

            if not op_name:
                continue

            def _get(*keys: str):
                for key in keys:
                    idx = header_map.get(key)
                    if idx is not None and idx < len(cells):
                        return cells[idx]
                return None

            time_seconds = _get("time_seconds", "sekundy", "czas[s]", "czas")
            setup_seconds = _get("setup_seconds", "przygotowanie[s]", "setup")
            work_center = _get("brygada", "workcenter", "work_center", "gniazdo")
            rate_value = _get("rate", "stawka")
            overhead_value = _get("overhead", "narzut")

            operations.append(
                RawOperationRow(
                    sourceFile=Path(path).name,
                    sheetName=sheet_name,
                    originalOperationName=op_name,
                    workCenter=str(work_center).strip() if work_center else None,
                    timeSeconds=float(time_seconds) if _is_number(time_seconds) else None,
                    setupTimeSeconds=float(setup_seconds) if _is_number(setup_seconds) else None,
                    ratePresent=bool(rate_value),
                    overheadPresent=bool(overhead_value),
                )
            )

    return {
        "sourceFile": Path(path).name,
        "detectedSheets": found_sheets,
        "operations": operations,
        "materialDetected": material_detected,
        "packagingDetected": packaging_detected,
        "costSummaryDetected": cost_summary_detected,
    }
